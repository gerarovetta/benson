import requests
import pandas as pd
import json
import csv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
import re


from http.server import BaseHTTPRequestHandler
from os.path import dirname, abspath, join
dir = dirname(abspath(__file__))
 
def render_chart(datasource, period, subfamily, user):
    url = "https://www.hinweiss.com/labs_services/renderChart"
    
    # Construct the sidebarData dynamically
   # sidebar_data_template = (
   #     "/product_sub_family_ds/value/datasource/{}/value/periodFilter/{}/grouping/0/"
   #     "product_model/product_model/filter/sub_family/{}/comparison/PREVIOUS_YEAR_CALENDAR/"
   #     "null/signature/LTzvv73vv70D77-9JCR-77-9du-_ve-_ve-_vU_vv73vv70z77-977-9B--_ve-_vUR5bu-_ve-_vTHvv73vv71s/"
   #     "sels/toolbar/chart/table/order/first/asc/extended/no/values/yes/toolbar/show"
   # )

    sidebar_data_template = (
       # "details/sels/datasource/{}/value/{}periodFilter/{}/grouping/0/product_model_color/product_model_color/filter/sub_family/{}/comparison/PREVIOUS_YEAR_CALENDAR/null/signature/eGzvv73vnZg877-977-977-9W0_vv73vv71E77-977-9Hu-_vQDvv70i77-9Y--_vVIQ77-977-9e--_vQ==/sels/toolbar/chart/table/order/first/asc/extended/no/values/yes/toolbar/show"
         "details/sels/datasource/{}/value/{}periodFilter/{}/grouping/0/product_model/product_model/filter/sub_family/{}/comparison/PREVIOUS_YEAR_CALENDAR/null/signature/eGzvv73vnZg877-977-977-9W0_vv73vv71E77-977-9Hu-_vQDvv70i77-9Y--_vVIQ77-977-9e--_vQ==/sels/toolbar/chart/table/order/first/asc/extended/no/values/yes/toolbar/show"
    )

    if user == "holkin" and datasource == "item_sales" :
        filterClient = "filter/client_category/Por+Mayor/"
    else:
        filterClient = ""
    
    
    sidebar_data = sidebar_data_template.format(datasource,filterClient,period, get_sub_family_id(subfamily,user))
    payload = json.dumps({
        "sidebarData": sidebar_data,
        "toolbarData": "toolbar/chart/table/order/first/asc/extended/yes/values/yes/toolbar/render"
    })
    if user == "holkin":
        play_session = f"0cdb85af12cc619b4673358eebb27f7a98828245-user={user}; Path=/; HTTPOnly; PLAY_SESSION=0cdb85af12cc619b4673358eebb27f7a98828245-user={user}"
    else:
        play_session = f"71186432b80de19be432135f406809e9c3a0af9b-user={user}; Path=/; HTTPOnly; PLAY_SESSION=71186432b80de19be432135f406809e9c3a0af9b-user={user}"


    
    headers = {
        'Cookie': play_session,
        'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)
    
    return response.text



def extract_table_data(json_response):
    """
    Extracts the header, rows, and footer from a JSON response.
    """
    data = json.loads(json_response)
    table = data.get("tables", [])[0]
    header = table.get("header", [])
    rows = table.get("data", [])
    footer = table.get("footer", [])

    return header, rows, footer




def expand_headers(headers):
    max_cols = sum(header['colspan'] for header in headers[0])
    max_rows = len(headers)
    table = [['' for _ in range(max_cols)] for _ in range(max_rows)]
    for row_idx, row in enumerate(headers):
        col_idx = 0
        for header in row:
            while table[row_idx][col_idx] != '':
                col_idx += 1
            for i in range(header.get('rowspan', 1)):
                for j in range(header['colspan']):
                    table[row_idx + i][col_idx + j] = header['data']
            col_idx += header['colspan']
    return table

def expand_data(data, max_cols):
    table = []
    for row in data:
        expanded_row = ['' for _ in range(max_cols)]
        col_idx = 0
        for cell in row:
            while expanded_row[col_idx] != '':
                col_idx += 1
            for j in range(cell['colspan']):
                cell_data = cell['data']
                # Check if the cell_data is numeric after removing commas
                if cell_data.replace('.', '', 1).isdigit():
                    expanded_row[col_idx + j] = int(cell_data.replace('.', '', 1))
                else:
                    expanded_row[col_idx + j] = cell['data']
            col_idx += cell['colspan']
        table.append(expanded_row)
    return table

def process_json(json_text):
    data = json.loads(json_text)
    header_data = data['tables'][0]['header']
    data_data = data['tables'][0]['data']
    footer_data = data['tables'][0].get('footer', [])
    expanded_headers = expand_headers(header_data)
    column_names = expanded_headers[-1]
    max_cols = len(column_names)
    expanded_data = expand_data(data_data, max_cols)
    expanded_footer = expand_data(footer_data, max_cols)
    combined_table = expanded_data
    return column_names, combined_table



def merge_and_process_responses(response_text_1, response_text_2, output_file, mergeComma):
    # Process the JSON responses into DataFrames
    column_names_1, table_1 = process_json(response_text_1)
    df1 = pd.DataFrame(table_1, columns=column_names_1)
    
    column_names_2, table_2 = process_json(response_text_2)
    df2 = pd.DataFrame(table_2, columns=column_names_2)

    # Merge the DataFrames based on the first column
    merged_df = pd.merge(df1, df2, on=column_names_1[0], how='outer')

    # Fill missing values with 0
    merged_df = merged_df.fillna(0)

    # Convert all columns except the first to numeric to ensure the sum operation works
    for col in merged_df.columns[1:]:
        if (mergeComma == ','):
            merged_df[col] = merged_df[col].astype(str).str.replace(mergeComma, '.').apply(lambda x: x.split()[0] if isinstance(x, str) and ' ' in x else x)
        else:
            merged_df[col] = merged_df[col].astype(str).str.replace(mergeComma, '').apply(lambda x: x.split()[0] if isinstance(x, str) and ' ' in x else x)
        merged_df[col] = pd.to_numeric(merged_df[col], errors='coerce').fillna(0)


    # Add a new column for the sum of each row
    merged_df['total'] = merged_df.iloc[:, 1:].sum(axis=1)

    # Save the merged table to a CSV file
    merged_df.to_csv(output_file, index=False)

    return merged_df

def merge_csv_exclude_first_column(csv1_path, csv2_path, output_path):
    # Read the first CSV file
    df1 = pd.read_csv(csv1_path)
    
    # Read the second CSV file
    df2 = pd.read_csv(csv2_path)
    
    # Drop the first column of the second CSV
    df2 = df2.drop(df2.columns[0], axis=1)
    
    # Append the second CSV to the first CSV
    merged_df = pd.concat([df1, df2], axis=1)
    
    # Save the merged DataFrame to the output CSV file
    merged_df.to_csv(output_path, index=False)

def merge_csv_with_first_column(csv1_path, csv2_path, output_path):
    # Read the first CSV file
    df1 = pd.read_csv(csv1_path)
    
    # Read the second CSV file
    df2 = pd.read_csv(csv2_path)
    
    # Iterate through the rows of df2
    for index, row in df2.iterrows():
        first_col_value = row.iloc[0]
        
        # Check if the first column value exists in df1
        if first_col_value in df1.iloc[:, 0].values:
            # If exists, update the corresponding row in df1 with the values from df2
            df1.loc[df1.iloc[:, 0] == first_col_value, df2.columns[1:]] = row.iloc[1:].values
        else:
            # If not exists, append the row to df1
            df1 = pd.concat([df1, row.to_frame().T], ignore_index=True)

    #total_row = pd.DataFrame(df1.sum(numeric_only=True)).T
    
    #total_row = total_row[df1.columns]  # Reorder columns to match original DataFrame
    
    # Concatenate the original DataFrame with the total row
    #df1 = pd.concat([df1, total_row], ignore_index=True)
    # Save the merged DataFrame to the output CSV file
    df1.to_csv(output_path, index=False)

def merge_csvs_on_title(csv1_path, csv2_path, output_path):
    # Read the first and second CSV files
    df1 = pd.read_csv(csv1_path)
    df2 = pd.read_csv(csv2_path)
    
    # Drop any columns with NaN values in df2
   # df2 = df2.dropna(axis=1)
    
    # Merge on the common column 'title'
    merged = df1.merge(df2, on='modelo', how='outer')
    
    # Save the merged DataFrame to the output CSV file
    merged.to_csv(output_path, index=False)

def merge_and_process_responses_new(subfamily, output_file):
    ventas_esta_temporada_benson = render_chart("item_sales", "THIS_SEASON", subfamily, "benson")
    ventas_esta_temporada_holkin = render_chart("item_sales", "THIS_SEASON", subfamily, "holkin")

    ventas_anterior_temporada_benson = render_chart("item_sales", "LAST_SEASON", subfamily, "benson")
    ventas_anterior_temporada_holkin = render_chart("item_sales", "LAST_SEASON", subfamily, "holkin")

    stock_benson = render_chart("stock_units", "THIS_SEASON", subfamily, "benson")
    stock_holkin = render_chart("stock_units", "THIS_SEASON", subfamily, "holkin")
    print("merge tests")
    print (stock_holkin)
    output_ventas_esta = f'{output_file}_ventas_esta.csv'
    output_ventas_ant = f'{output_file}_ventas_ant.csv'
    output_stock = f'{output_file}_stock.csv'
    output_path = f'{output_file}_merged_output.csv'
    output_final = f'{output_file}'

    # Process each set of data and save to CSV
    merge_and_process_responses(ventas_esta_temporada_benson, ventas_esta_temporada_holkin, output_ventas_esta,",")
    merge_and_process_responses(ventas_anterior_temporada_benson, ventas_anterior_temporada_holkin, output_ventas_ant,",")
    merge_and_process_responses(stock_benson, stock_holkin, output_stock,".")
    print(output_stock)
    # Merge ventas_ant and ventas_esta
    df_merged = merge_csvs_on_title(output_ventas_ant, output_ventas_esta, output_path)
    # Merge output_path and stock
    df_final = merge_csvs_on_title(output_path, output_stock, output_final)

    return df_final

def add_totals_to_csv(input_file, output_file):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(input_file)
    df = df.rename(columns={df.columns[1]: 'N-UnidadesVendidas Anterior'})
    df = df.rename(columns={df.columns[2]: 'H-UnidadesVendidas Anterior'})
    df = df.rename(columns={df.columns[4]: 'N-UnidadesVendidas Actual'})
    df = df.rename(columns={df.columns[5]: 'H-UnidadesVendidas Actual'})
    df = df.rename(columns={df.columns[7]: 'Nekil Unidades Stock'})
    df = df.rename(columns={df.columns[8]: 'Holkin Unidades Stock'})
    # Compute the totals for each column (excluding non-numeric columns)
    totals = df.select_dtypes(include='number').sum()
    
    # Convert totals to a DataFrame and transpose it to match the column structure
    totals_df = pd.DataFrame(totals).transpose()
    
    # Rename the index to 'Total'
    totals_df.index = ['Total']
    
    # Append the totals row to the original DataFrame
    df_with_totals = pd.concat([df, totals_df], ignore_index=True)
    # Save the modified DataFrame to a new CSV file
    df_with_totals.to_csv(output_file, index=False)
    
    return df_with_totals

def process_subfamilies(subfamilies):
    output_files = []
    self.wfile.write(str('Processing subfamilies').encode())
    for subfamily_ in subfamilies:
        self.wfile.write(str('Processing subfamilie 1').encode())
        print(f'Processing {subfamily_}')
        output_file = 'merged_data.csv'
        self.wfile.write(str('Processing subfamilie 2').encode())
        processed_df = merge_and_process_responses_new(subfamily_, output_file)
        self.wfile.write(str('Processing subfamilie 3').encode())
        processedFile = f'{subfamily_}.csv'
        self.wfile.write(str('Processing subfamilie 4').encode())
        add_totals_to_csv(output_file,processedFile)
        self.wfile.write(str('Processing subfamilie 5').encode())
        output_files.append(processedFile)
        print(f'Merged and processed data saved to: {output_file}_merged_total_subfamily.csv')
        self.wfile.write(str('Processing subfamilie 6').encode())

    return output_files

def csvs_to_excel_with_formatting(csv_filenames, output_excel_file):
    """
    Converts a list of CSV files into an Excel file with each CSV as a separate sheet and applies specific formatting.

    Parameters:
    csv_filenames (list of str): List of paths to CSV files.
    output_excel_file (str): Path to the output Excel file.
    """
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:

        # Define the border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for csv_filename in csv_filenames:
            # Read each CSV file into a DataFrame
            df = pd.read_csv(csv_filename)
            # Extract the name of the file (without extension) to use as the sheet name
            sheet_name = csv_filename.split('/')[-1].split('.')[0]
            # Write the DataFrame to a sheet
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Access the workbook and sheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Set the width of the first column
            worksheet.column_dimensions['A'].width = 30
            
            # Add headers for "Temporada Actual" and "Temporada Anterior"
            worksheet.insert_rows(1)
            worksheet.insert_rows(1)
            worksheet.merge_cells('B1:D1')
            worksheet['B1'] = 'Temporada Anterior'
            worksheet['B2'] = 'Nekil'
            worksheet['C2'] = 'Holkin'
            worksheet['B1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('E1:G1')
            worksheet['E1'] = 'Temporada Actual'
            worksheet['E2'] = 'Nekil'
            worksheet['F2'] = 'Holkin'
            worksheet['E1'].alignment = Alignment(horizontal='center')
            worksheet.merge_cells('H1:J1')
            worksheet['H1'] = 'Stock'
            worksheet['H2'] = 'Nekil'
            worksheet['I2'] = 'Holkin'
            worksheet['H1'].alignment = Alignment(horizontal='center')
            
            # Apply color fill to the columns
            fill_colors = [
                'EFFFE0', 'E0FFCC', 'E6E6FA', # Light colors
                'EFFACD', 'D8BFD8', 'FFDAB9', # Light colors
                'E0FFCC' # Light colors
            ]
            
            yellow_tones = ['FFFFE0', 'FFFACD', 'FFEFD5']  # Tonos amarillos
            red_tones = ['FFC0CB', 'FFB6C1', 'FF69B4']     # Tonos rojos
            blue_tones = ['ADD8E6', 'B0E0E6', '87CEEB']   

            for i, col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H']):
                for cell in worksheet[col]:
                    cell.fill = PatternFill(start_color=fill_colors[i % len(fill_colors)], end_color=fill_colors[i % len(fill_colors)], fill_type='solid')
            
            #first column
            for i, col in enumerate(['A']):
                for cell in worksheet[col]:
                    cell.fill = PatternFill(start_color=fill_colors[i % len(fill_colors)], end_color=fill_colors[i % len(fill_colors)], fill_type='solid')
            
            for i, col in enumerate(['B','C', 'D']):
                for cell in worksheet[col]:
                    cell.fill = PatternFill(start_color=yellow_tones[i % len(yellow_tones)], end_color=yellow_tones[i % len(yellow_tones)], fill_type='solid')
            
            for i, col in enumerate(['E','F', 'G']):
                for cell in worksheet[col]:
                    cell.fill = PatternFill(start_color=red_tones[i % len(red_tones)], end_color=red_tones[i % len(red_tones)], fill_type='solid')
            
            for i, col in enumerate(['H','I', 'J']):
                for cell in worksheet[col]:
                    cell.fill = PatternFill(start_color=blue_tones[i % len(blue_tones)], end_color=blue_tones[i % len(blue_tones)], fill_type='solid')
            # Apply bold font to the headers
            for cell in worksheet["1:1"]:
                cell.font = Font(bold=True)
            
            for i,col in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H','I']):
                adjusted_width = 25
                worksheet.column_dimensions[col].width = adjusted_width
            
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            worksheet.freeze_panes = worksheet['B4']

#csvs_to_excel_with_formatting(output_csv_files,'NekilHolkin.xlsx')


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-type', 'text/plain')
        self.end_headers()
        self.wfile.write(str('Hello World 2!!').encode())
        self.wfile.write(str('Hello World 3!!').encode())
        self.wfile.write(str('Hello World 4!!').encode())
        output_csv_files = process_subfamilies(subfamilies)
        self.wfile.write(str('Hello World 4!!').encode())
        for file in output_csv_files:
            self.wfile.write(str('Hello World 2!!').encode())
        self.wfile.write(str('Hello World!!').encode())
        csvs_to_excel_with_formatting(output_csv_files,'NekilHolkin.xlsx')
        
        self.wfile.write(str('Hello end!!').encode())
        return